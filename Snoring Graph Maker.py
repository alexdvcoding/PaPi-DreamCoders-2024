import time
import matplotlib.pyplot as plt
from openpyxl import Workbook
from datetime import datetime, timedelta
import random as rand
import RPi.GPIO as GPIO
import time
import random
SOUND_SENSOR_PIN=7

GPIO.setmode(GPIO.BCM)
GPIO.setup(SOUND_SENSOR_PIN, GPIO.IN)
prev_sound_state = GPIO.input(SOUND_SENSOR_PIN)



# This is a placeholder for your sensor reading function
def read_sensor():
    global prev_sound_state
    amplitude = 0
    for i in range(20):
        sound_state=GPIO.input(SOUND_SENSOR_PIN)
            
        if sound_state!= prev_sound_state:
            if sound_state == GPIO.LOW:
                amplitude+=240               
        prev_sound_state = sound_state
        time.sleep(0.02)
    return amplitude

# Create a new workbook and select the active sheet
wb = Workbook()
ws = wb.active

# Record data for 8 hours (480 minutes)
end_time = datetime.now() + timedelta(seconds=15)

# Lists to store time and amplitude data for plotting
time_data = []
amplitude_data = []

while datetime.now() < end_time:
    # Read sensor data
    amplitude = read_sensor()

    # Write the time and amplitude level to the spreadsheet
    ws.append([datetime.now(), amplitude])

    # Save the workbook
    wb.save("sensor_data.xlsx")

    # Add the data to our lists
    time_data.append(datetime.now())
    amplitude_data.append(amplitude)

    # Wait for 60 seconds
    time.sleep(0)

# At the end of the 8 hours, plot the data
plt.figure(figsize=(10, 6))  # Adjust the size of the figure
plt.plot(time_data, amplitude_data, marker='o')  # Add markers for each data point
plt.xlabel('Time')
plt.ylabel('Amplitude')
plt.title('Amplitude Level Over Time')
plt.grid(True)

# Set the y-axis limits and ticks
plt.ylim(0, 1000)
plt.yticks(range(0, 1001, 100))

# Ensure that the y-axis labels match the data points
plt.gca().set_yticklabels([str(int(i)) for i in plt.gca().get_yticks()])

plt.tight_layout()  # Adjust the layout to ensure all elements fit in the figure
plt.savefig('sensor_data.png', dpi=300)  # Save the figure with a higher resolution

# At the end of the day, reset the spreadsheet by creating a new workbook
if datetime.now().hour == 0:
    wb = Workbook()
    ws = wb.active
    time_data = []
    amplitude_data = []
reset=0
if reset==0:
    wb = Workbook()
    ws = wb.active
    time_data = []
    amplitude_data = [] 
print('ra')